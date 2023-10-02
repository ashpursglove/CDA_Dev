import psychrolib
import openpyxl
import math
import tkinter as tk
from tkinter import messagebox

# Constants
CO2_LATENT_HEAT = 571.08  # kJ/kg
H2O_LATENT_HEAT = 2260  # kJ/kg
CO2_LIX_DENSITY = 1.101  # kg/L
H2O_MR = 18.02  # g/mol
CO2_MR = 44.01  # g/mol

# Initialize psychrolib
psychrolib.SetUnitSystem(psychrolib.SI)


def calculate_wet_bulb_temperature(dry_bulb_temperature, relative_humidity, pressure):
    """
    Calculates the wet-bulb temperature based on the dry-bulb temperature, relative humidity, and pressure.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in degrees Celsius.
        relative_humidity (float): Relative humidity as a percentage.
        pressure (float): Atmospheric pressure in Pa.

    Returns:
        float: Wet-bulb temperature in degrees Celsius.
    """
    # Convert relative humidity from percentage to decimal
    relative_humidity_decimal = relative_humidity / 100.0

    # Calculate the wet-bulb temperature
    wet_bulb_temperature = psychrolib.GetTWetBulbFromRelHum(dry_bulb_temperature, relative_humidity_decimal, pressure)

    return wet_bulb_temperature


def calculate_humidity_ratio(dry_bulb_temperature, relative_humidity, pressure):
    """
    Calculates the humidity ratio based on dry-bulb temperature, relative humidity, and pressure.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in degrees Celsius.
        relative_humidity (float): Relative humidity as a percentage.
        pressure (float): Atmospheric pressure in Pa.

    Returns:
        float: Humidity ratio in kilograms of water vapor per kilogram of dry air.
    """
    # Convert relative humidity from percentage to decimal
    relative_humidity_decimal = relative_humidity / 100.0

    # Calculate the humidity ratio
    humidity_ratio = psychrolib.GetHumRatioFromRelHum(dry_bulb_temperature, relative_humidity_decimal, pressure)

    return humidity_ratio


def calculate_partial_vapor_pressure(dry_bulb_temperature, relative_humidity):
    """
    Calculates the partial vapor pressure based on dry-bulb temperature and relative humidity.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in °C [SI].
        relative_humidity (float): Relative humidity as a percentage.

    Returns:
        float: Partial vapor pressure in Pascals.
    """
    # Convert relative humidity from percentage to decimal
    relative_humidity_decimal = relative_humidity / 100.0

    # Calculate the partial vapor pressure
    partial_pressure = psychrolib.GetVapPresFromRelHum(dry_bulb_temperature, relative_humidity_decimal)

    return partial_pressure


def calculate_dry_air_enthalpy(dry_bulb_temperature):
    """
    Calculates the dry air enthalpy based on temperature.

    Args:
        dry_bulb_temperature (float): Temperature in degrees Celsius.

    Returns:
        float: Dry air enthalpy in joules per kilogram of dry air.
    """
    # Calculate the dry air enthalpy
    dry_air_enthalpy = psychrolib.GetDryAirEnthalpy(dry_bulb_temperature)

    return dry_air_enthalpy


def calculate_wet_air_enthalpy(dry_bulb_temperature, humidity_ratio):
    """
    Calculates the wet air enthalpy based on dry-bulb temperature and humidity ratio.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in degrees Celsius.
        humidity_ratio (float): Humidity ratio in kilograms of water vapor per kilogram of dry air.

    Returns:
        float: Wet air enthalpy in joules per kilogram of moist air.
    """
    # Calculate the wet air enthalpy
    wet_air_enthalpy = psychrolib.GetMoistAirEnthalpy(dry_bulb_temperature, humidity_ratio)

    return wet_air_enthalpy


def calculate_air_density(dry_bulb_temperature, humidity_ratio, pressure):
    """
    Calculates the air density based on dry-bulb temperature, humidity ratio, and pressure.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in degrees Celsius.
        humidity_ratio (float): Humidity ratio in kilograms of water vapor per kilogram of dry air.
        pressure (float): Atmospheric pressure in Pascals.

    Returns:
        float: Air density in kilograms per cubic meter.
    """
    # Calculate the air density
    air_density = psychrolib.GetMoistAirDensity(dry_bulb_temperature, humidity_ratio, pressure)

    return air_density


def calculate_psychrometrics(dry_bulb_temperature, wet_bulb_temperature, pressure):
    """
    Calculates various psychrometric properties based on dry-bulb temperature, wet-bulb temperature, and pressure.

    Args:
        dry_bulb_temperature (float): Dry-bulb temperature in degrees Celsius.
        wet_bulb_temperature (float): Wet-bulb temperature in degrees Celsius.
        pressure (float): Atmospheric pressure in Pascals.

    Returns:
        tuple: A tuple containing the humidity ratio, dew-point temperature, relative humidity,
               partial pressure of water vapor, moist air enthalpy, specific volume of moist air,
               and degree of saturation.
    """
    # Calculate psychrometric properties
    result = psychrolib.CalcPsychrometricsFromTWetBulb(dry_bulb_temperature, wet_bulb_temperature, pressure)

    return result


def calculate_pressure_from_altitude(altitude):
    """
    Calculates the atmospheric pressure based on altitude.

    Args:
        altitude (float): Altitude in meters.

    Returns:
        float: Atmospheric pressure in Pascals.
    """
    # Calculate the atmospheric pressure
    pressure = psychrolib.GetStandardAtmPressure(altitude)

    return pressure



def read_excel_data(file_path, col, start_row, num_of_rows):
    """
    Reads data from an Excel sheet and retrieves a column of data from a given position.

    Args:
        file_path (str): Path to the Excel file.
        col (int): Starting column position.
        start_row (int): Starting row position.
        num_of_rows (int): Number of rows to read.

    Returns:
        list: A list of values from the specified column. Non-numeric values are replaced with 0.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Initialize an empty list to store the data
    data_list = []

    # Loop through the rows and extract the data
    for row in range(num_of_rows):
        value = sheet.cell(row + start_row, column=col).value
        try:
            numeric_value = float(value)
            data_list.append(numeric_value)
        except (ValueError, TypeError):
            data_list.append(0)

    # Close the workbook
    workbook.close()

    # Return the data list
    return data_list


def get_altitude_and_airflow(file_path):
    """
    Reads data from an Excel sheet and retrieves altitude and air flow.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        tuple: A tuple containing the altitude in meters, air flow in litres per second.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Retrieve altitude and air flow values
    altitude = float(sheet.cell(row=5, column=3).value)
    air_flow = float(sheet.cell(row=6, column=3).value)
    resin_diameter = float(sheet.cell(row=5, column=7).value)
    resin_mass = float(sheet.cell(row=6, column=7).value)
    resin_density = float(sheet.cell(row=7, column=7).value)
    chamber_diameter = float(sheet.cell(row=9, column=7).value)

    # Close the workbook
    workbook.close()

    # Return the data
    return altitude, air_flow, resin_diameter, resin_mass, resin_density, chamber_diameter


def calculate_energy_flux(moist_air_enthalpy, mass_flow):
    """
    Calculates the energy flux based on moist air enthalpy and mass flow.

    Args:
        moist_air_enthalpy (float): Moist air enthalpy in joules per kilogram of moist air.
        mass_flow (float): Mass flow rate in kilograms per second.

    Returns:
        float: Energy flux in joules per second.
    """
    # Calculate the energy flux
    energy_flux = moist_air_enthalpy * mass_flow

    return energy_flux


def calculate_co2_flow(co2_level, air_flow):
    """
    Calculates the CO2 flow based on CO2 level and air flow.

    Args:
        co2_level (float): CO2 level in ppm.
        air_flow (float): Air flow rate in liters per second.

    Returns:
        float: CO2 flow rate in milligrams per second.
    """
    # Calculate the CO2 flow rate
    co2_flow = co2_level * air_flow

    return co2_flow


def calculate_co2_change(co2_flow_inlet, co2_flow_outlet):
    """
    Calculates the change in CO2 flow.

    Args:
        co2_flow_inlet (float): CO2 flow rate at the inlet in milligrams per second.
        co2_flow_outlet (float): CO2 flow rate at the outlet in milligrams per second.

    Returns:
        float: Change in CO2 flow rate in milligrams per second.
    """
    # Calculate the change in CO2 flow rate
    co2_change = co2_flow_outlet - co2_flow_inlet

    return co2_change


def calculate_co2_change_label(co2_change):
    """
    Determines the label for the change in CO2 flow.

    Args:
        co2_change (float): Change in CO2 flow rate in milligrams per second.

    Returns:
        str: Label for the change in CO2 flow.
    """
    if co2_change > 0:
        return "CO2 Release"
    elif co2_change < 0:
        return "CO2 Capture"
    else:
        return "No Change in CO2"


def process_file():
    # Get the file path from the entry widget
    file_path = entry_file_path.get()

    try:
        # Append ".xlsx" to the file path
        file_path += ".xlsx"

        # Get the data from the file
        altitude, air_flow, resin_diameter, resin_mass, resin_density, chamber_diameter = get_altitude_and_airflow(file_path)
        pressure = calculate_pressure_from_altitude(altitude)

        # Calculate surface area and volume of a single sphere
        resin_diameter_m = resin_diameter / 1000  # Convert resin diameter from mm to m
        resin_radius = resin_diameter_m / 2
        single_sphere_surface_area = math.pi * resin_diameter_m ** 2
        single_sphere_volume = (4 / 3) * math.pi * resin_radius ** 3

        # Calculate total resin volume
        total_resin_volume = 1 / (resin_density / resin_mass)

        # Calculate the rough number of spheres
        rough_number_of_spheres = total_resin_volume / single_sphere_volume

        # Calculate the total surface area
        total_surface_area = rough_number_of_spheres * single_sphere_surface_area

        # First data set (column 3, row 8 to row 10)
        dry_bulb_temperature_inlet, relative_humidity_inlet, co2_level_inlet = read_excel_data(file_path, 3, 8, 3)
        wet_bulb_temperature_inlet = calculate_wet_bulb_temperature(dry_bulb_temperature_inlet, relative_humidity_inlet,
                                                                    pressure)
        result_inlet = calculate_psychrometrics(dry_bulb_temperature_inlet, wet_bulb_temperature_inlet, pressure)
        humidity_ratio_inlet, dew_point_temperature_inlet, calc_relative_humidity_inlet, partial_pressure_inlet, \
        moist_air_enthalpy_inlet, specific_volume_inlet, degree_of_saturation_inlet = result_inlet
        dry_air_enthalpy_inlet = calculate_dry_air_enthalpy(dry_bulb_temperature_inlet)
        air_density_inlet = calculate_air_density(dry_bulb_temperature_inlet, humidity_ratio_inlet, pressure)

        # Calculate CO2 flow for inlet data set
        co2_flow_inlet = calculate_co2_flow(co2_level_inlet, air_flow)

        # Second data set (column 3, row 12 to row 14)
        dry_bulb_temperature_outlet, relative_humidity_outlet, co2_level_outlet = read_excel_data(file_path, 3, 12, 3)
        wet_bulb_temperature_outlet = calculate_wet_bulb_temperature(dry_bulb_temperature_outlet,
                                                                     relative_humidity_outlet, pressure)
        result_outlet = calculate_psychrometrics(dry_bulb_temperature_outlet, wet_bulb_temperature_outlet, pressure)
        humidity_ratio_outlet, dew_point_temperature_outlet, calc_relative_humidity_outlet, partial_pressure_outlet, \
        moist_air_enthalpy_outlet, specific_volume_outlet, degree_of_saturation_outlet = result_outlet
        dry_air_enthalpy_outlet = calculate_dry_air_enthalpy(dry_bulb_temperature_outlet)
        air_density_outlet = calculate_air_density(dry_bulb_temperature_outlet, humidity_ratio_outlet, pressure)

        # Calculate CO2 flow for outlet data set
        co2_flow_outlet = calculate_co2_flow(co2_level_outlet, air_flow)

        # Calculate the change in CO2
        co2_change = calculate_co2_change(co2_flow_inlet, co2_flow_outlet)

        # Determine the label for CO2 change
        co2_change_label = calculate_co2_change_label(co2_change)

        # Calculate energy flux for inlet and outlet
        mass_flow = air_density_outlet * (air_flow / 1000)  # Convert air flow from L/s to m³/s
        energy_flux_inlet = calculate_energy_flux(moist_air_enthalpy_inlet, mass_flow)
        energy_flux_outlet = calculate_energy_flux(moist_air_enthalpy_outlet, mass_flow)

        # Calculate arease and volumes for chamber

        c_radius = chamber_diameter / 2
        chamber_area = math.pi * (c_radius ** 2)
        c_area_m2 = chamber_area/10000
        air_m3_s = air_flow / 1000
        chamber_speed = air_m3_s / c_area_m2



        # Prepare the output message
        output_message = ""
        output_message += "Altitude: {} m\n".format(altitude)
        output_message += "Air Flow: {} L/s\n".format(air_flow)
        output_message += "Air Flow: {} m³/s\n".format(air_m3_s)
        output_message += "\n"
        output_message += "Chamber Diameter: {} mm\n".format(chamber_diameter*10)
        output_message += "Chamber Area: {} cm²\n".format(chamber_area)
        output_message += "Chamber Area: {} m²\n".format(c_area_m2)
        output_message += "Gas Speed in Chamber: {} m/s\n".format(chamber_speed)
        output_message += "Gas Speed in Chamber: {} cm/s\n".format(chamber_speed*100)
        output_message += "\n"
        output_message += "Resin Diameter: {} mm\n".format(resin_diameter)
        output_message += "Resin Mass: {} kg\n".format(resin_mass)
        output_message += "Resin Density: {} kg/m³\n".format(resin_density)
        output_message += "Single Sphere Surface Area: {} m²\n".format(single_sphere_surface_area)
        output_message += "Single Sphere Volume: {} m³\n".format(single_sphere_volume)
        output_message += "Total Resin Volume: {} m³\n".format(total_resin_volume)
        output_message += "Rough Number of Spheres: {}\n".format(rough_number_of_spheres)
        output_message += "Total Surface Area: {} m²\n".format(total_surface_area)
        output_message += "Mass Flow: {} kg/s\n".format(mass_flow)
        output_message += "\n"
        output_message += "Inlet Data Set:\n"
        output_message += "Dry Bulb Temperature: {} °C\n".format(dry_bulb_temperature_inlet)
        output_message += "Relative Humidity: {} %\n".format(relative_humidity_inlet)
        output_message += "CO2 Level: {}\n".format(co2_level_inlet)
        output_message += "Wet Bulb Temperature: {} °C\n".format(wet_bulb_temperature_inlet)
        output_message += "Humidity Ratio: {} kg/kg\n".format(humidity_ratio_inlet)
        output_message += "Dew Point Temperature: {} °C\n".format(dew_point_temperature_inlet)
        output_message += "Relative Humidity (Calculated): {} %\n".format(calc_relative_humidity_inlet)
        output_message += "Partial Pressure: {} Pa\n".format(partial_pressure_inlet)
        output_message += "Moist Air Enthalpy: {} J/kg\n".format(moist_air_enthalpy_inlet)
        output_message += "Specific Volume: {} m³/kg\n".format(specific_volume_inlet)
        output_message += "Degree of Saturation: {}\n".format(degree_of_saturation_inlet)
        output_message += "Dry Air Enthalpy: {} J/kg\n".format(dry_air_enthalpy_inlet)
        output_message += "Air Density: {} kg/m³\n".format(air_density_inlet)
        output_message += "CO2 Flow (Inlet): {} mg/s\n".format(co2_flow_inlet)
        output_message += "Energy Flux: {} J/s\n".format(energy_flux_inlet)
        output_message += "\n"
        output_message += "Outlet Data Set:\n"
        output_message += "Dry Bulb Temperature: {} °C\n".format(dry_bulb_temperature_outlet)
        output_message += "Relative Humidity: {} %\n".format(relative_humidity_outlet)
        output_message += "CO2 Level: {}\n".format(co2_level_outlet)
        output_message += "Wet Bulb Temperature: {} °C\n".format(wet_bulb_temperature_outlet)
        output_message += "Humidity Ratio: {} kg/kg\n".format(humidity_ratio_outlet)
        output_message += "Dew Point Temperature: {} °C\n".format(dew_point_temperature_outlet)
        output_message += "Relative Humidity (Calculated): {} %\n".format(calc_relative_humidity_outlet)
        output_message += "Partial Pressure: {} Pa\n".format(partial_pressure_outlet)
        output_message += "Moist Air Enthalpy: {} J/kg\n".format(moist_air_enthalpy_outlet)
        output_message += "Specific Volume: {} m³/kg\n".format(specific_volume_outlet)
        output_message += "Degree of Saturation: {}\n".format(degree_of_saturation_outlet)
        output_message += "Dry Air Enthalpy: {} J/kg\n".format(dry_air_enthalpy_outlet)
        output_message += "Air Density: {} kg/m³\n".format(air_density_outlet)
        output_message += "CO2 Flow (Outlet): {} mg/s\n".format(co2_flow_outlet)
        output_message += "Energy Flux: {} J/s\n".format(energy_flux_outlet)
        output_message += "\n"
        output_message += "Changes:\n"
        output_message += "Dry Bulb Temperature Change: {} °C\n".format(dry_bulb_temperature_outlet - dry_bulb_temperature_inlet)
        output_message += "Relative Humidity Change: {} %\n".format(relative_humidity_outlet - relative_humidity_inlet)
        output_message += "CO2 Level Change: {}\n".format(co2_level_outlet - co2_level_inlet)
        output_message += "Wet Bulb Temperature Change: {} °C\n".format(wet_bulb_temperature_outlet - wet_bulb_temperature_inlet)
        output_message += "Humidity Ratio Change: {} kg/kg\n".format(humidity_ratio_outlet - humidity_ratio_inlet)
        output_message += "Dew Point Temperature Change: {} °C\n".format(dew_point_temperature_outlet - dew_point_temperature_inlet)
        output_message += "Relative Humidity (Calculated) Change: {} %\n".format(calc_relative_humidity_outlet - calc_relative_humidity_inlet)
        output_message += "Partial Pressure Change: {} Pa\n".format(partial_pressure_outlet - partial_pressure_inlet)
        output_message += "Moist Air Enthalpy Change: {} J/kg\n".format(moist_air_enthalpy_outlet - moist_air_enthalpy_inlet)
        output_message += "Specific Volume Change: {} m³/kg\n".format(specific_volume_outlet - specific_volume_inlet)
        output_message += "Degree of Saturation Change: {}\n".format(degree_of_saturation_outlet - degree_of_saturation_inlet)
        output_message += "Dry Air Enthalpy Change: {} J/kg\n".format(dry_air_enthalpy_outlet - dry_air_enthalpy_inlet)
        output_message += "Air Density Change: {} kg/m³\n".format(air_density_outlet - air_density_inlet)
        output_message += "{}: {} mg/s\n".format(co2_change_label, co2_change)

        # Display the output message in the text widget
        text_output.delete(1.0, tk.END)
        text_output.insert(tk.END, output_message)
    except:
        pass


# Create the main window
window = tk.Tk()
window.title("Ash's Rough Draft")
window.geometry("1000x1200")
window.configure(bg="#e8a023")

# Create the file path entry widget
label_file_path = tk.Label(window, text="Enter File Name:", bg="#f2f2f2")
label_file_path.pack()

entry_file_path = tk.Entry(window, width=50)
entry_file_path.pack()

# Create the process button
button_process = tk.Button(window, text="Process File", command=process_file, bg="#6fa8dc", fg="white")
button_process.pack()

# Create the output text widget
text_output = tk.Text(window, height=70, width=80)
text_output.pack()

# Create the exit button
def exit_program():
    window.destroy()

button_exit = tk.Button(window, text="Exit", command=exit_program, bg="red", fg="white")
button_exit.place(relx=1, rely=1, anchor="se")

# Start the main loop
window.mainloop()